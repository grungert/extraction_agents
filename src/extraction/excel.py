"""Excel file extraction utilities for Excel Header Mapper."""
import os
import tempfile
import pandas as pd
from decimal import Decimal
from markitdown import MarkItDown
from ..utils.display import console, logger # Import logger
from ..models import AppConfig
from rich.panel import Panel
import re
import openpyxl
import xlrd

# Import custom exceptions (Corrected import path)
from src.exceptions import FileProcessingError # Import FileProcessingError

CURRENCY_REGEX = re.compile(r'\[\$([^\-\]]+)(?:\-[0-9A-F]+)?\]')

def format_value_based_on_excel_format(value, num_format):
    if value is None:
        return ""

    is_numeric = isinstance(value, (int, float, Decimal))

    if not is_numeric or not num_format or num_format == 'General':
        return str(value)

    currency_match = CURRENCY_REGEX.search(num_format)
    if currency_match:
        currency_symbol = currency_match.group(1)
        decimal_places = 2 if '.00' in num_format else 0
        try:
            num_decimal = Decimal(str(value))
            formatted_num = "{:,.{dp}f}".format(num_decimal, dp=decimal_places)
            if num_format.startswith('['):
                return f"{currency_symbol}{formatted_num}"
            else:
                return f"{formatted_num} {currency_symbol}"
        except Exception:
            return str(value)

    if '%' in num_format:
        try:
            num_decimal = Decimal(str(value)) * 100
            decimal_places = 0
            if '.0%' in num_format:
                decimal_places = 1
            elif '.00%' in num_format:
                decimal_places = 2
            formatted_num = "{:,.{dp}f}".format(num_decimal, dp=decimal_places)
            return f"{formatted_num}%"
        except Exception:
            return str(value)

    try:
        num_decimal = Decimal(str(value))
        decimal_places = 2 if '.' in str(value) else 0
        if '.00' in num_format: decimal_places = 2
        if '.000' in num_format: decimal_places = 3
        return "{:,.{dp}f}".format(num_decimal, dp=decimal_places)
    except Exception:
        return str(value)

def get_formatted_dataframe(file_path, sheet_name, start_row, end_row):
    ext = os.path.splitext(file_path)[-1].lower()
    nrows = end_row - start_row

    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            rows = list(ws.iter_rows(min_row=start_row + 1, max_row=end_row, values_only=False))
            data = []
            for row in rows:
                formatted_row = []
                for cell in row:
                    value = cell.value
                    num_format = cell.number_format
                    formatted_value = format_value_based_on_excel_format(value, num_format)
                    formatted_row.append(formatted_value)
                data.append(formatted_row)

            return pd.DataFrame(data)
        except FileNotFoundError:
            raise FileProcessingError(f"File not found: {file_path}")
        except Exception as e:
            raise FileProcessingError(f"Error reading XLSX file {file_path}: {e}")


    elif ext == ".xls":
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            xf_list = wb.xf_list
            fmt_map = wb.format_map
            data = []

            for r_idx in range(start_row, min(end_row, sheet.nrows)):
                row_data = []
                for c_idx in range(sheet.ncols):
                    cell = sheet.cell(r_idx, c_idx)
                    val = cell.value
                    fmt = None
                    if cell.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_DATE):
                        xf_idx = cell.xf_index
                        if xf_idx < len(xf_list):
                            fmt_key = xf_list[xf_idx].format_key
                            fmt_obj = fmt_map.get(fmt_key)
                            if fmt_obj:
                                fmt = fmt_obj.format_str
                    formatted = format_value_based_on_excel_format(val, fmt)
                    row_data.append(formatted)
                data.append(row_data)

            return pd.DataFrame(data)
        except FileNotFoundError:
            raise FileProcessingError(f"File not found: {file_path}")
        except Exception as e:
            raise FileProcessingError(f"Error reading XLS file {file_path}: {e}")

    else:
        # Raise custom FileProcessingError for unsupported format
        raise FileProcessingError(f"Unsupported file format: {ext}. Please provide a .csv, .xls, or .xlsx file.")

def excel_to_markdown(file_path, config: AppConfig, sheet_name=None):
    try:
        # Use lowercase extension checks for case-insensitive matching
        file_path_lower = file_path.lower()
        nrows = config.end_row - config.start_row

        if file_path_lower.endswith('.csv'):
            from pathlib import Path
            from csv_to_md import csv_to_markdown
            rows = config.end_row - config.start_row
            markdown = csv_to_markdown(Path(file_path), rows)
            # Remove header and separator rows from markdown output
            lines = markdown.splitlines()
            if len(lines) > 2:
                markdown = "\n".join(lines[2:])
            logger.info(f"Converted CSV to markdown (rows {config.start_row}-{config.end_row-1})") # Use logger.info
            return markdown

        elif file_path_lower.endswith(('.xls', '.xlsx')):
            sheet_param = sheet_name if sheet_name else (0 if not config.all_sheets else None)
            # get_formatted_dataframe now raises FileProcessingError
            df = get_formatted_dataframe(file_path, sheet_param, config.start_row, config.end_row)
            temp_suffix = '.xlsx'

            with tempfile.NamedTemporaryFile(suffix=temp_suffix, delete=False) as temp_file:
                temp_path = temp_file.name
            df.to_excel(temp_path, index=False, header=False, engine='openpyxl')

        else:
            # This case should ideally be caught by get_formatted_dataframe, but kept for safety
            raise FileProcessingError(f"Unsupported file format: {file_path_lower}. Please provide a .csv, .xls, or .xlsx file.")

        md = MarkItDown(enable_plugins=False, nrows=nrows)
        result = md.convert(temp_path)
        os.unlink(temp_path)

        logger.info(f"Converted to markdown (rows {config.start_row}-{config.end_row-1})") # Use logger.info
        return result.text_content

    except FileProcessingError:
        # Re-raise custom exception
        raise
    except Exception as e:
        # Catch other unexpected errors during markdown conversion
        logger.exception(f"Error converting file to markdown: {str(e)}") # Use logger.exception
        raise FileProcessingError(f"Error converting file to markdown: {str(e)}")


def prepare_excel_sheets_markdown(file_path, config: AppConfig):
    sheet_markdowns = {}

    try:
        if file_path.lower().endswith('.csv'):
            # excel_to_markdown now raises FileProcessingError
            markdown = excel_to_markdown(file_path, config)
            if markdown:
                logger.info(f"Processed CSV file to markdown (rows {config.start_row} to {config.end_row-1})") # Use logger.info
                sheet_markdowns["csv_data"] = markdown
            else:
                logger.warning("No data found in CSV file") # Use logger.warning

        elif file_path.lower().endswith(('.xls', '.xlsx')):
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names

            logger.info(f"Converting {len(sheet_names)} sheets to markdown...") # Use logger.info

            for idx, sheet in enumerate(sheet_names):
                logger.info(f"Sheet {idx + 1}/{len(sheet_names)}: {sheet}") # Use logger.info
                # get_formatted_dataframe now raises FileProcessingError
                df = get_formatted_dataframe(file_path, sheet, config.start_row, config.end_row)

                if df.empty:
                    logger.warning(f"Sheet '{sheet}' is empty - skipping") # Use logger.warning
                    continue

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                    temp_path = temp_file.name
                df.to_excel(temp_path, index=False, header=False, engine='openpyxl')

                md = MarkItDown(enable_plugins=False, nrows=config.end_row - config.start_row)
                result = md.convert(temp_path)
                sheet_markdowns[sheet] = result.text_content
                os.unlink(temp_path)

                logger.info(f"Converted sheet '{sheet}' to markdown") # Use logger.info

        else:
            # Raise custom FileProcessingError for unsupported format
            raise FileProcessingError(f"Unsupported file format: {file_path}. Please provide a .csv, .xls, or .xlsx file.")

    except FileProcessingError:
        # Re-raise custom exception
        raise
    except Exception as e:
        # Catch other unexpected errors during sheet processing
        logger.exception(f"Error preparing excel sheets to markdown: {str(e)}") # Use logger.exception
        raise FileProcessingError(f"Error preparing excel sheets to markdown: {str(e)}")
