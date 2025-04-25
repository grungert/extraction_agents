import openpyxl
import os
import xlrd
import win32com.client

def convert_xls2xlsx_excel(file_path_xls, file_path_xlsx=None):
    """
    Converts an .xls file to .xlsx using Microsoft Excel automation (Windows only).

    Args:
        file_path_xls (str): The path to the input .xls file.
        file_path_xlsx (str, optional): The path to save the .xlsx file.
                                        If not specified, it will be saved in the same
                                        location with the extension changed.
    """
    try:
        excel = win32com.client.Dispatch("Excel.Application")               # Create an instance of the Excel application object
        workbook = excel.Workbooks.Open(file_path_xls)                      # Open the .xls workbook
        # If the output path for .xlsx is not provided
        if file_path_xlsx is None:
            # Get the base name and extension of the input file
            base, ext = os.path.splitext(file_path_xls)
            # Create the output path by changing the extension to .xlsx
            file_path_xlsx = base + ".xlsx"

        # Save the workbook in .xlsx format (FileFormat=51)
        workbook.SaveAs(file_path_xlsx)
        # Close the workbook
        workbook.Close()
        # Quit the Excel application
        excel.Quit()

        print(f"The file '{file_path_xls}' has been saved as '{file_path_xlsx}'")

    except FileNotFoundError:
        print(f"Error: The file '{file_path_xls}' was not found.")
    except Exception as e:
        print(f"An error occurred during the conversion (Is Microsoft Excel installed?): {e}")

def convert_xls2xlsx_using_xlrd_and_openpyxl(file_path_xls, file_path_xlsx=None):
    try:
        workbook_xls = xlrd.open_workbook(file_path_xls)

        if file_path_xlsx is None:
            base, ext = os.path.splitext(file_path_xls)
            file_path_xlsx = base + "_1.xlsx"

        workbook_xlsx = openpyxl.Workbook()

        for sheet_name in workbook_xls.sheet_names():
            sheet_xls = workbook_xls.sheet_by_name(sheet_name)
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_name)

            for row_idx in range(sheet_xls.nrows):
                for col_idx in range(sheet_xls.ncols):
                    cell_value = sheet_xls.cell_value(row_idx, col_idx)
                    sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        if 'Sheet' in workbook_xlsx.sheetnames and len(workbook_xlsx.sheetnames) > 1:
            del workbook_xlsx['Sheet']
        elif len(workbook_xlsx.sheetnames) == 1 and workbook_xlsx.sheetnames[0] == 'Sheet' and not workbook_xls.sheet_names():
            workbook_xlsx.remove(workbook_xlsx['Sheet'])
            workbook_xlsx.create_sheet(title='Sheet1') # Create a default sheet if the xls was empty

        workbook_xlsx.save(file_path_xlsx)
        print(f"'{file_path_xls}' saved as '{file_path_xlsx}'.")

    except FileNotFoundError:
        print(f"Error: File not found.")
    except Exception as e:
        print(f"An error occurred while converting the .xls file: {e}")

def get_column_formats_xlsx(file_path, sheet_name=None):
    try:
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                print(f"Warning: Sheet '{sheet_name}' not found. Using the first sheet.")
                sheet = workbook.active
        else:
            sheet = workbook.active  # Default to the first sheet

        formats = {}
        max_col = sheet.max_column

        # Iterate through each column
        for col_idx in range(1, max_col + 1):
            # Get the format for the cell in row 378 of this column
            cell_ = sheet.cell(row=1, column=col_idx)
            format_string = cell_.number_format
            formats[col_idx - 1] = format_string  # Adjust index to be 0-based

        return formats

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
        return {}
    except Exception as e:
        print(f"An error occurred while reading the .xlsx file: {e}")
        return {}
    
def get_column_formats_xls_enhanced(file_path, sheet_name=None):
    try:
        workbook = xlrd.open_workbook(file_path)
        if sheet_name:
            sheet = workbook.sheet_by_name(sheet_name)
        else:
            sheet = workbook.sheet_by_index(0)

        formats = {}
        num_cols = sheet.ncols

        for col_idx in range(num_cols):
            cell_ = sheet.cell(378, col_idx)
            format_string = None

            # Check for cell-level formatting
            if cell_.xf_index is not None:
                format_key = workbook.xf_list[cell_.xf_index].format_key
                format_string = workbook.format_map[format_key].format_str
            else:
                print(f"Cell ({378}, {col_idx}) has no specific formatting information. Checking column default.")
                try:
                    colinfo = sheet.colinfo_map.get(col_idx)
                    if colinfo and colinfo.xf_index is not None:
                        format_key = workbook.xf_list[colinfo.xf_index].format_key
                        format_string = workbook.format_map[format_key].format_str
                        print(f"Found column default format: {format_string}")
                    else:
                        print(f"Column {col_idx} also has no specific default formatting.")
                except AttributeError:
                    print(f"Warning: Could not access column info for column {col_idx}.")

            formats[col_idx] = format_string

        return formats

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
        return {}
    except Exception as e:
        print(f"An error occurred while reading the .xls file: {e}")
        return {}

file = 'C:/Users/SuneyToste/Flows_Suney/8206743_COUPONS pour externe.2025.01.27.xls'
name_doc, extension =  os.path.splitext(file)
# if extension == '.xls':
    # convert_xls2xlsx_using_xlrd_and_openpyxl(file)
    # convert_xls2xlsx_excel(file)
column_formats = get_column_formats_xls_enhanced(name_doc+'.xls')
print(column_formats)