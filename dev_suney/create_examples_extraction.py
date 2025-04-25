#!/usr/bin/env python3
# /// script
# dependencies = [
#   "rich>=13.7.0",
#   "pandas>=2.0.0",
#   "xlrd >= 2.0.1"
# ]
# ///
import pandas as pd
import chardet
import lxml.etree as ET
from io import StringIO
import json
import os
import random
from rich.console import Console
from rich.table import Table
from rich import print_json
import xlrd
import openpyxl

def detect_delimiter(file_path: str, num_lines: int = 5) -> str:
    """ Attempts to detect the delimiter (comma or semicolon) in a .dat file.
    Args: file_path (str): The path to the .dat file.
        num_lines (int, optional): The number of lines to inspect. Defaults to 5.
    Returns: str: The detected delimiter ("," or ";"), "Undetermined" if frequencies are similar,
                  "No delimiters found" if neither comma nor semicolon is found in the inspected lines,
                 or an error message if the file is not found."""
    try:
        with open(file_path, 'r') as file:
            comma_count = 0
            semicolon_count = 0
            lines_read = 0
            for _ in range(num_lines):
                line = file.readline()
                if not line:
                    break  # End of file reached
                comma_count += line.count(',')
                semicolon_count += line.count(';')
                lines_read += 1
            if comma_count > semicolon_count:
                return ","
            elif semicolon_count > comma_count:
                return ";"
            elif comma_count > 0 or semicolon_count > 0:
                return "Undetermined"  # Similar frequencies
            else:
                return "No delimiters found in the first {} lines.".format(lines_read)
    except FileNotFoundError:
        return f"Error: File '{file_path}' not found."
    

def xlrd_workbook_to_dataframe_all_sheets(workbook: xlrd.book.Book) -> dict[str, pd.DataFrame]:
    """Converts all sheets of an xlrd workbook object into a dictionary of Pandas DataFrames.
    Args: workbook (xlrd.book.Book): The xlrd workbook object.
    Returns: dict[str, pandas.DataFrame]: A dictionary where keys are sheet names and values are the corresponding DataFrames."""
    all_sheets_data = {}
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        data = []
        header = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        for row_num in range(1, sheet.nrows):
            row_values = []
            for col_num in range(sheet.ncols):
                cell_value = sheet.cell_value(row_num, col_num)
                row_values.append(cell_value)
            data.append(row_values)
        df = pd.DataFrame(data, columns=header)
        all_sheets_data[sheet_name] = df
    return all_sheets_data

def openpyxl_workbook_to_dataframe_all_sheets(workbook: openpyxl.workbook.workbook.Workbook) -> dict[str, pd.DataFrame]:
    """Converts all sheets of an openpyxl workbook object into a dictionary of Pandas DataFrames.
    Args: workbook (openpyxl.workbook.workbook.Workbook): The openpyxl workbook object.
    Returns: dict[str, pandas.DataFrame]: A dictionary where keys are sheet names and values are the corresponding DataFrames. """
    all_sheets_data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        header = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        df = pd.DataFrame(data, columns=header)
        all_sheets_data[sheet_name] = df
    return all_sheets_data

console = Console()

# Format values to avoid scientific notation and clean dates
def clean_value(x):
    if isinstance(x, float):
        return f"{x:.4f}".rstrip('0').rstrip('.')
    if pd.isna(x):
        return " "
    if isinstance(x, pd.Timestamp):
        return x.strftime("%Y-%m-%d")
    return str(x)

def loading_data_as_df(extension: str, document_file_path):        
    if extension.lower() in ('.csv','.txt'):
        separator = detect_delimiter(document_file_path, num_lines = 5) 
        with open(document_file_path, 'rb') as f:
            data_crude = f.read()
            guessing = chardet.detect(data_crude)
            encode2use = guessing['encoding']
        try:
            if separator in (",", ";"):
                df_all = pd.read_csv(document_file_path, encoding=encode2use, header=None, sep= separator)
            else:
                try:
                    df_all = pd.read_csv(document_file_path, encoding=encode2use, header=None, sep= ',') 
                except Exception as e:
                    print(f"Error trying to open the document with pandas: {e}")
        except Exception as e:
            print(f"Error trying to open the document with pandas: {e}")
    if extension.lower() == '.dat':
            separator = detect_delimiter(document_file_path, num_lines = 5) 
            try:
                if separator in (",", ";"):
                    df_all = pd.read_csv(document_file_path, header=None, sep=',')
                else:
                    try:
                        df_all = pd.read_csv(document_file_path, encoding=encode2use, header=None, sep= ',')
                    except Exception as e:
                        print(f"Error trying to open the document with pandas as separeted by , or ; : {e}")
            except Exception as e:
                    print(f"Error trying to open the document with pandas as separeted by , or ; : {e}")
    if extension.lower() in ('.xls', '.xlsx') :
            try: 
                try:          
                    df_all = pd.read_excel(document_file_path, header=None)
                except Exception as e:
                    if extension.lower() == '.xlsx':
                        wb = openpyxl.load_workbook(document_file_path)
                        data = openpyxl_workbook_to_dataframe_all_sheets(wb)
                        if len(data.keys())==1:
                            df_all = data[data.keys()[0]]
                    elif extension.lower() == '.xls':
                        wb = xlrd.open_workbook(document_file_path)
                        data = xlrd_workbook_to_dataframe_all_sheets(wb)
                        if len(data.keys())==1:
                            df_all = data[data.keys()[0]]
            except Exception as e:
                print(f"Error trying to open the document with pandas: {e}")
    if extension.lower() == '.xml':
            try:
                parser = ET.XMLParser(recover=True)
                tree = ET.parse(document_file_path, parser=parser)
                root = tree.getroot()
                xml_string = ET.tostring(root, encoding='utf-8').decode('utf-8')
                df_all = pd.read_xml(StringIO(xml_string))
            except Exception as e:
                print(f"Error trying to open the document with pandas: {e}")
    if extension.lower() == '.html':
            try:
                df_all = pd.read_html(document_file_path)
            except Exception as e:
                print(f"Error trying to open the document with pandas: {e}")
    
    #Removing columns and rows that are fully empty
    empty_columns = df_all.columns[df_all.isna().all()]
    df_all = df_all.drop(empty_columns, axis=1)
    empty_rows = df_all.isna().all(axis=1)
    df_all = df_all[~empty_rows]

    # Clean all values
    df_all = df_all.map(clean_value)
    return df_all

# # File paths
# document_file_path = 'Fluxs - AI Training/8216969_Format#17.2025.01.30'
document_file_path = 'C:/Users/SuneyToste/Flows_Suney/8217081_FIDOEIC7.2025.01.30.CSV'
# json_file_path = 'ground-truth/file_17_H.json'

# # Load JSON metadata
# with open(json_file_path, 'r') as f:
#     metadata = json.load(f)

# # Extract metadata
# header_start = metadata["Context"]["HeaderStartLine"]
# header_end = metadata["Context"]["HeaderEndLine"]
# content_start = metadata["Context"]["ContentStartLine"]

# # CHANGE THIS TO THE CORRECT MODEL
# json_output_fields = metadata["Identifier"][0]  # assume first object in list
# # json_output_fields = metadata["Denomination"]
# # json_output_fields = metadata["Valorisation"]
# # json_output_fields = metadata["MarketCap"]
# # json_output_fields = metadata["CorporateAction"][0]
# # json_output_fields = metadata["Characteristics"]

# # Load entire document to access specific rows
name_doc, extension =  os.path.splitext(document_file_path)

df_all = loading_data_as_df(extension, document_file_path)

# # Slice rows to include headers and 5 rows of content
# if extension.lower() in ('.csv', '.xls', '.xlsx'):
#     header_start = header_start - 1
#     header_end = header_end - 1
#     content_start  = content_start - 1
# rows_needed = list(range(header_start, header_end + 1)) + list(range(content_start, content_start + min(5, df_all.shape[0]-content_start)))
# df_slice = df_all.iloc[rows_needed].reset_index(drop=True)

# # Try to detect column indexes based on the last header row
# num_header_rows = header_end - header_start + 1
# header_row = df_slice.iloc[num_header_rows-1]
# column_map = {}
# for key, col_name in json_output_fields.items():
#     if col_name:
#         for idx, val in enumerate(header_row):
#             if str(val).strip().lower() == col_name.strip().lower():
#                 column_map[key] = idx
#                 break

# # Clean all values
# df_slice_clean = df_slice.map(clean_value)

# # Extract and clean header row
# temp_header_df = pd.DataFrame(columns=df_slice_clean.columns)
# for j in range(num_header_rows):
#     row_cleaned = df_slice_clean.iloc[j].map(str).fillna(" ").map(str.strip)
#     temp_header_df.loc[j] = row_cleaned
# header_values = []
# for col in temp_header_df.columns:
#     columns_values = [' ' if element == 'nan' else element for element in temp_header_df[col].tolist()]
#     header_values.append(' '.join(columns_values))
# header_row_cleaned = pd.Series(header_values)

# # Assign headers and drop the original header row
# df_slice_clean.columns = header_row_cleaned
# df_slice_clean = df_slice_clean.drop(index = list(range(0, num_header_rows))).reset_index(drop=True)

# # Re-filter from cleaned dataframe using updated headers
# # Filter dataframe to only selected columns
# filtered_cols = list(column_map.values())
# filtered_cols_clean = [col for i, col in enumerate(header_row_cleaned) if i in filtered_cols]
# df_filtered = df_slice_clean[filtered_cols_clean]

# # Generate clean markdown
# markdown_table = df_slice_clean.to_markdown(index=False)

# # Output metadata
# ValidationConfidence = round(random.uniform(0.90, 1.00), 2)
# json_output_fields["ValidationConfidence"] = ValidationConfidence

# # === Markdown Output ===
# print("\n# Original Table (Markdown)")
# print(markdown_table)

# # JSON result
# console.rule("[cyan]JSON Output with Confidence")
# print_json(data=json_output_fields)

# # === Rich Output ===
# console.rule("[bold blue]Rich Output")
# # Input table (with real headers)
# console.rule("[green]Input Table Slice (Full)")
# input_table = Table(show_lines=True)
# nb_col = 0
# for col in df_slice_clean.columns:
#     input_table.add_column(str(col))
#     nb_col += 1
# print(f"Number columns: {nb_col}")
# for _, row in df_slice_clean.iterrows():
#     row_str = [str(x) for x in row]
#     print(f"Number of elements in row: {len(row_str)}")
#     input_table.add_row(*row_str)
# console.print(input_table)

# console.rule("[green]Input Filtered Table Slice (Full)")
# input_table2 = Table(show_lines=True)
# for col in df_filtered.columns:
#     input_table2.add_column(str(col))
# for _, row in df_filtered.iterrows():
#     input_table2.add_row(*[str(x) for x in row])
# console.print(input_table2)


